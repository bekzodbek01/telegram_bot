from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from pathlib import Path
from datetime import datetime
from db import (
    all_staff_with_stats,
    get_stats,
    list_staff,
    get_month_stats
)

EXPORT_DIR = Path("exports")
EXPORT_DIR.mkdir(exist_ok=True)


# ======================= SET COLUMN WIDTHS =======================

def set_column_widths(ws):
    widths = {
        1: 7,     # ID
        2: 13,     # Name
        3: 30,     # Position
        4: 13,     # Region
        5: 6,     # Likes
        6: 6,     # Dislikes
        7: 6,     # Neutrals
        8: 6,     # Total
        9: 10,     # Date
        10: 10,    # Month
        11: 6     # Year
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


# ======================= STYLE HEADER =======================

def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


# ======================= EXPORT ALL STAFF =======================

def export_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "All Staff"

    ws.append([
        "ID", "Name", "Position", "Region",
        "Likes", "Dislikes", "Neutrals", "Total",
        "Date", "Month", "Year"
    ])

    set_column_widths(ws)
    style_header(ws)

    now = datetime.now()
    date_str = now.strftime("%d.%m.%Y")
    month_name = now.strftime("%B")
    year = now.year

    rows = all_staff_with_stats()

    for item in rows:
        ws.append([
            item["staff_id"],
            item["name"],
            item["position"],
            item["region"],
            item["likes"],
            item["dislikes"],
            item["neutrals"],
            item["total"],
            date_str,
            month_name,
            year
        ])

    filename = EXPORT_DIR / "all_staff_report.xlsx"
    wb.save(filename)
    return str(filename)


# ======================= EXPORT ONE STAFF =======================

def export_one_staff_excel(staff_id: int):
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Report"

    ws.append([
        "ID", "Name", "Position", "Region",
        "Likes", "Dislikes", "Neutrals", "Total",
        "Date", "Month", "Year"
    ])

    set_column_widths(ws)
    style_header(ws)

    now = datetime.now()
    date_str = now.strftime("%d.%m.%Y")
    month_name = now.strftime("%B")
    year = now.year

    staff_info = None
    for sid, name, pos, reg in list_staff():
        if sid == staff_id:
            staff_info = {"name": name, "position": pos, "region": reg}
            break

    if not staff_info:
        raise ValueError(f"Staff with ID {staff_id} not found!")

    stats = get_stats(staff_id)

    ws.append([
        staff_id,
        staff_info["name"],
        staff_info["position"],
        staff_info["region"],
        stats["likes"],
        stats["dislikes"],
        stats["neutrals"],
        stats["total"],
        date_str,
        month_name,
        year
    ])

    filename = EXPORT_DIR / f"staff_{staff_id}_report.xlsx"
    wb.save(filename)
    return str(filename)


# ======================= EXPORT MONTH/YEAR REPORT =======================

def export_month_excel(year: int, month: int):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}_{year}"

    ws.append([
        "ID", "Name", "Position", "Region",
        "Likes", "Dislikes", "Neutrals", "Total",
        "Month", "Year"
    ])

    set_column_widths(ws)
    style_header(ws)

    month_stats = get_month_stats(year, month)

    staff_dict = {
        sid: {"name": n, "position": p, "region": r}
        for sid, n, p, r in list_staff()
    }

    for item in month_stats:
        sid = item["staff_id"]
        if sid not in staff_dict:
            continue

        ws.append([
            sid,
            staff_dict[sid]["name"],
            staff_dict[sid]["position"],
            staff_dict[sid]["region"],
            item["likes"],
            item["dislikes"],
            item["neutrals"],
            item["total"],
            month,
            year
        ])

    filename = EXPORT_DIR / f"month_{month}_{year}.xlsx"
    wb.save(filename)
    return str(filename)
